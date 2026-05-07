"""
Reports API endpoints - Professional Executive PDF Report
"""
from io import BytesIO
from uuid import UUID
from collections import Counter
from datetime import datetime

from fastapi import APIRouter, Depends, HTTPException, status
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from sqlalchemy.orm import Session
from sqlalchemy import func

from app.api.deps import get_current_user
from app.core.database import get_db
from app.models.audit import Audit
from app.models.finding import Finding, FindingStatus
from app.models.sap import SAPUser
from app.models.sod import Conflict, SoDRule, RuleSeverity
from app.models.user import AppUser


router = APIRouter(prefix="/audits/{audit_id}/reports", tags=["Reports"])


def _get_audit_or_404(db: Session, audit_id: UUID) -> Audit:
    audit = db.query(Audit).filter(Audit.id == audit_id).first()
    if not audit:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Audit not found")
    return audit


# ──────────────────────────────────────────────
#  Excel Export (unchanged)
# ──────────────────────────────────────────────
@router.get("/conflicts.xlsx")
def export_conflicts_excel(
    audit_id: UUID,
    db: Session = Depends(get_db),
    current_user: AppUser = Depends(get_current_user),
):
    _get_audit_or_404(db, audit_id)
    rows = db.query(
        Conflict, SAPUser.user_id, SAPUser.full_name, SoDRule.name, SoDRule.severity,
    ).join(SAPUser, Conflict.sap_user_id == SAPUser.id
    ).join(SoDRule, Conflict.rule_id == SoDRule.id
    ).filter(Conflict.audit_id == audit_id).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Conflicts"

    headers = ["Conflict ID", "SAP User", "Full Name", "Rule", "Severity",
               "Risk Score", "Detected At", "Set A TCodes", "Set B TCodes"]
    header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="0B3B68", end_color="0B3B68", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    for row_idx, (conflict, user_id, full_name, rule_name, severity) in enumerate(rows, 2):
        values = [
            str(conflict.id), user_id, full_name or "", rule_name,
            severity.value, conflict.risk_score, conflict.detected_at.isoformat(),
            ", ".join(conflict.tcodes_set_a or []), ", ".join(conflict.tcodes_set_b or []),
        ]
        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = thin_border

    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="conflicts_{audit_id}.xlsx"'},
    )


# ──────────────────────────────────────────────
#  Professional Executive PDF Report
# ──────────────────────────────────────────────
@router.get("/executive.pdf")
def export_executive_pdf(
    audit_id: UUID,
    db: Session = Depends(get_db),
    current_user: AppUser = Depends(get_current_user),
):
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.units import mm, cm
        from reportlab.lib.colors import HexColor, white, black
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_JUSTIFY, TA_RIGHT
        from reportlab.platypus import (
            SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle,
            PageBreak, KeepTogether, HRFlowable,
        )
        from reportlab.graphics.shapes import Drawing, String, Rect, Line, Circle
        from reportlab.graphics.charts.piecharts import Pie
        from reportlab.graphics.charts.barcharts import VerticalBarChart
        from reportlab.graphics.widgets.markers import makeMarker
    except ImportError:
        raise HTTPException(
            status_code=status.HTTP_503_SERVICE_UNAVAILABLE,
            detail="reportlab is required. pip install reportlab",
        )

    # ── 1. Fetch all data ──────────────────────
    audit = _get_audit_or_404(db, audit_id)
    total_users = db.query(SAPUser).filter(SAPUser.audit_id == audit_id).count()
    total_conflicts = db.query(Conflict).filter(Conflict.audit_id == audit_id).count()
    findings = db.query(Finding).filter(Finding.audit_id == audit_id).all()
    open_findings = len([f for f in findings if f.status.value in ("OPEN", "IN_REVIEW")])

    # Severity distribution
    severity_counts = dict(
        db.query(SoDRule.severity, func.count(Conflict.id))
        .join(Conflict, Conflict.rule_id == SoDRule.id)
        .filter(Conflict.audit_id == audit_id)
        .group_by(SoDRule.severity).all()
    )
    high_count = severity_counts.get(RuleSeverity.HIGH, 0)
    medium_count = severity_counts.get(RuleSeverity.MEDIUM, 0)
    low_count = severity_counts.get(RuleSeverity.LOW, 0)

    # Finding status distribution
    finding_status_counts = Counter(f.status.value for f in findings)

    # Top conflicts by rule
    conflicts_per_rule = (
        db.query(SoDRule.name, func.count(Conflict.id).label("cnt"))
        .join(Conflict, Conflict.rule_id == SoDRule.id)
        .filter(Conflict.audit_id == audit_id)
        .group_by(SoDRule.name)
        .order_by(func.count(Conflict.id).desc())
        .limit(8).all()
    )

    # Top 20 critical conflicts
    critical_conflicts = (
        db.query(Conflict, SAPUser.user_id, SAPUser.full_name, SoDRule.name, SoDRule.severity)
        .join(SAPUser, Conflict.sap_user_id == SAPUser.id)
        .join(SoDRule, Conflict.rule_id == SoDRule.id)
        .filter(Conflict.audit_id == audit_id)
        .order_by(Conflict.risk_score.desc())
        .limit(20).all()
    )

    # Users with most conflicts
    top_users = (
        db.query(SAPUser.user_id, SAPUser.full_name, func.count(Conflict.id).label("cnt"))
        .join(Conflict, Conflict.sap_user_id == SAPUser.id)
        .filter(Conflict.audit_id == audit_id)
        .group_by(SAPUser.user_id, SAPUser.full_name)
        .order_by(func.count(Conflict.id).desc())
        .limit(10).all()
    )

    # ── 2. Define colors & styles ──────────────
    PRIMARY = HexColor("#0B3B68")
    ACCENT = HexColor("#1E88E5")
    RED = HexColor("#E53935")
    ORANGE = HexColor("#FB8C00")
    GREEN = HexColor("#43A047")
    LIGHT_BG = HexColor("#F5F7FA")
    DARK_TEXT = HexColor("#1A1A2E")
    GRAY = HexColor("#6B7280")
    LIGHT_GRAY = HexColor("#E5E7EB")
    WHITE = white

    styles = getSampleStyleSheet()
    styles.add(ParagraphStyle(
        "ReportTitle", fontName="Helvetica-Bold", fontSize=22, textColor=PRIMARY,
        spaceAfter=4, alignment=TA_LEFT,
    ))
    styles.add(ParagraphStyle(
        "ReportSubtitle", fontName="Helvetica", fontSize=11, textColor=GRAY,
        spaceAfter=12, alignment=TA_LEFT,
    ))
    styles.add(ParagraphStyle(
        "SectionTitle", fontName="Helvetica-Bold", fontSize=14, textColor=PRIMARY,
        spaceBefore=18, spaceAfter=8, borderPadding=(0, 0, 4, 0),
    ))
    styles.add(ParagraphStyle(
        "BodyText2", fontName="Helvetica", fontSize=9.5, textColor=DARK_TEXT,
        spaceAfter=6, alignment=TA_JUSTIFY, leading=13,
    ))
    styles.add(ParagraphStyle(
        "SmallGray", fontName="Helvetica", fontSize=8, textColor=GRAY, alignment=TA_LEFT,
    ))
    styles.add(ParagraphStyle(
        "KPIValue", fontName="Helvetica-Bold", fontSize=24, textColor=PRIMARY,
        alignment=TA_CENTER, spaceAfter=2,
    ))
    styles.add(ParagraphStyle(
        "KPILabel", fontName="Helvetica", fontSize=8, textColor=GRAY,
        alignment=TA_CENTER, spaceAfter=0,
    ))
    styles.add(ParagraphStyle(
        "TableHeader", fontName="Helvetica-Bold", fontSize=8.5, textColor=WHITE,
        alignment=TA_CENTER,
    ))
    styles.add(ParagraphStyle(
        "TableCell", fontName="Helvetica", fontSize=8, textColor=DARK_TEXT,
        alignment=TA_LEFT,
    ))
    styles.add(ParagraphStyle(
        "TableCellCenter", fontName="Helvetica", fontSize=8, textColor=DARK_TEXT,
        alignment=TA_CENTER,
    ))

    # ── 3. Build PDF elements ──────────────────
    elements = []
    page_w, page_h = A4
    content_w = page_w - 50 * 2  # 50pt margins

    # ── COVER / HEADER ─────────────────────────
    # Title block with colored line
    elements.append(Spacer(1, 10))
    elements.append(HRFlowable(width="100%", thickness=3, color=PRIMARY, spaceAfter=12))
    elements.append(Paragraph("REPORTE EJECUTIVO DE AUDITORIA", styles["ReportTitle"]))
    elements.append(Paragraph("Segregacion de Funciones (SoD) en SAP", styles["ReportSubtitle"]))
    elements.append(HRFlowable(width="100%", thickness=1, color=LIGHT_GRAY, spaceAfter=16))

    # Audit info table (clean two-column layout)
    info_data = [
        [Paragraph("<b>Auditoria:</b>", styles["BodyText2"]),
         Paragraph(f"{audit.name}", styles["BodyText2"])],
        [Paragraph("<b>Empresa:</b>", styles["BodyText2"]),
         Paragraph(f"{audit.company_name}", styles["BodyText2"])],
        [Paragraph("<b>Periodo:</b>", styles["BodyText2"]),
         Paragraph(f"{audit.period_start} a {audit.period_end}", styles["BodyText2"])],
        [Paragraph("<b>Fecha del reporte:</b>", styles["BodyText2"]),
         Paragraph(f"{datetime.now().strftime('%d/%m/%Y %H:%M')}", styles["BodyText2"])],
    ]
    info_table = Table(info_data, colWidths=[content_w * 0.3, content_w * 0.7])
    info_table.setStyle(TableStyle([
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
        ("TOPPADDING", (0, 0), (-1, -1), 3),
    ]))
    elements.append(info_table)
    elements.append(Spacer(1, 16))

    # ── SECTION 1: KPI Cards ──────────────────
    elements.append(Paragraph("1. RESUMEN EJECUTIVO", styles["SectionTitle"]))
    elements.append(Paragraph(
        "A continuacion se presenta un resumen de los principales indicadores de la auditoria "
        "de Segregacion de Funciones (SoD) realizada sobre los usuarios del sistema SAP. "
        "La auditoria evalua si existen usuarios con accesos que permitan ejecutar actividades "
        "incompatibles, lo cual representa un riesgo de fraude o error.",
        styles["BodyText2"],
    ))
    elements.append(Spacer(1, 8))

    # Risk level
    if total_conflicts == 0:
        risk_text = "BAJO"
        risk_color = GREEN
    elif high_count > total_conflicts * 0.3:
        risk_text = "CRITICO"
        risk_color = RED
    elif high_count > 0:
        risk_text = "ALTO"
        risk_color = ORANGE
    else:
        risk_text = "MODERADO"
        risk_color = ORANGE

    conflict_pct = round((total_conflicts / max(total_users, 1)) * 100, 1)

    def make_kpi_cell(value, label, color=PRIMARY):
        return [
            Paragraph(f'<font color="{color.hexval()}">{value}</font>', styles["KPIValue"]),
            Paragraph(label, styles["KPILabel"]),
        ]

    kpi_data = [[
        make_kpi_cell(str(total_users), "Usuarios SAP\nevaluados"),
        make_kpi_cell(str(total_conflicts), "Conflictos SoD\ndetectados", RED if total_conflicts > 0 else GREEN),
        make_kpi_cell(str(open_findings), "Hallazgos\nabiertos"),
        make_kpi_cell(risk_text, "Nivel de\nriesgo global", risk_color),
    ]]

    # Flatten: each cell is a list of 2 Paragraphs → need to make it a nested table
    kpi_inner = []
    for cell_content in kpi_data[0]:
        inner = Table([[cell_content[0]], [cell_content[1]]], colWidths=[content_w / 4 - 8])
        inner.setStyle(TableStyle([
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("TOPPADDING", (0, 0), (-1, -1), 8),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
            ("BACKGROUND", (0, 0), (-1, -1), LIGHT_BG),
            ("BOX", (0, 0), (-1, -1), 0.5, LIGHT_GRAY),
            ("ROUNDEDCORNERS", [4, 4, 4, 4]),
        ]))
        kpi_inner.append(inner)

    kpi_table = Table([kpi_inner], colWidths=[content_w / 4] * 4)
    kpi_table.setStyle(TableStyle([
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
    ]))
    elements.append(kpi_table)
    elements.append(Spacer(1, 6))

    elements.append(Paragraph(
        f"<b>Interpretacion:</b> Del total de <b>{total_users}</b> usuarios SAP evaluados, "
        f"se detectaron <b>{total_conflicts}</b> conflictos de segregacion de funciones, "
        f"lo que representa un <b>{conflict_pct}%</b> de incidencia. "
        f"El nivel de riesgo global se clasifica como <b>{risk_text}</b>.",
        styles["BodyText2"],
    ))

    # ── SECTION 2: Donut Chart - Severity ──────
    elements.append(Spacer(1, 8))
    elements.append(Paragraph("2. DISTRIBUCION DE CONFLICTOS POR SEVERIDAD", styles["SectionTitle"]))
    elements.append(Paragraph(
        "Los conflictos detectados se clasifican en tres niveles de severidad: "
        "<b>ALTO</b> (riesgo critico de fraude), <b>MEDIO</b> (riesgo moderado que requiere atencion) "
        "y <b>BAJO</b> (riesgo menor, usualmente informativo). El siguiente grafico muestra "
        "la proporcion de conflictos en cada nivel.",
        styles["BodyText2"],
    ))

    if total_conflicts > 0:
        donut_drawing = Drawing(content_w, 180)

        pie = Pie()
        pie.x = content_w / 2 - 70
        pie.y = 20
        pie.width = 140
        pie.height = 140
        pie.data = [high_count or 0.001, medium_count or 0.001, low_count or 0.001]
        pie.labels = None
        pie.slices.strokeWidth = 0.5
        pie.slices.strokeColor = WHITE
        # Donut effect

        severity_colors = [RED, ORANGE, GREEN]
        for i, color in enumerate(severity_colors):
            pie.slices[i].fillColor = color
            pie.slices[i].popout = 2

        donut_drawing.add(pie)
        donut_drawing.add(Circle(
            pie.x + (pie.width / 2),
            pie.y + (pie.height / 2),
            (min(pie.width, pie.height) / 2) * 0.55,
            fillColor=WHITE,
            strokeColor=WHITE,
        ))

        # Legend
        legend_x = content_w / 2 + 100
        legend_items = [
            (RED, f"ALTO: {high_count} ({round(high_count/max(total_conflicts,1)*100)}%)"),
            (ORANGE, f"MEDIO: {medium_count} ({round(medium_count/max(total_conflicts,1)*100)}%)"),
            (GREEN, f"BAJO: {low_count} ({round(low_count/max(total_conflicts,1)*100)}%)"),
        ]
        for i, (color, text) in enumerate(legend_items):
            y_pos = 130 - i * 22
            donut_drawing.add(Rect(legend_x, y_pos, 12, 12, fillColor=color, strokeColor=None))
            donut_drawing.add(String(legend_x + 18, y_pos + 2, text,
                                     fontName="Helvetica", fontSize=9, fillColor=DARK_TEXT))

        elements.append(donut_drawing)
    else:
        elements.append(Paragraph("<i>No se detectaron conflictos en esta auditoria.</i>", styles["BodyText2"]))

    # ── SECTION 3: Bar Chart - Conflicts per Rule ──
    if conflicts_per_rule:
        elements.append(Spacer(1, 8))
        elements.append(Paragraph("3. CONFLICTOS POR REGLA DE SoD", styles["SectionTitle"]))
        elements.append(Paragraph(
            "El siguiente grafico de barras muestra cuantos conflictos fueron detectados por cada "
            "regla de Segregacion de Funciones. Las reglas con mayor cantidad de conflictos son las "
            "que requieren atencion prioritaria, ya que afectan a mas usuarios.",
            styles["BodyText2"],
        ))

        bar_drawing = Drawing(content_w, 200)
        bc = VerticalBarChart()
        bc.x = 60
        bc.y = 30
        bc.width = content_w - 120
        bc.height = 150
        bc.data = [[cnt for _, cnt in conflicts_per_rule]]
        bc.categoryAxis.categoryNames = [name[:25] for name, _ in conflicts_per_rule]
        bc.categoryAxis.labels.fontName = "Helvetica"
        bc.categoryAxis.labels.fontSize = 7
        bc.categoryAxis.labels.angle = 30
        bc.categoryAxis.labels.boxAnchor = "ne"
        bc.valueAxis.valueMin = 0
        bc.valueAxis.labels.fontName = "Helvetica"
        bc.valueAxis.labels.fontSize = 8
        bc.bars[0].fillColor = ACCENT
        bc.bars[0].strokeColor = None
        bc.barWidth = 18
        bc.groupSpacing = 8

        bar_drawing.add(bc)
        elements.append(bar_drawing)

    # ── SECTION 4: Top 20 Critical Conflicts Table ──
    section_num = 4 if conflicts_per_rule else 3
    elements.append(Spacer(1, 8))
    elements.append(Paragraph(f"{section_num}. TOP 20 CONFLICTOS DE MAYOR RIESGO", styles["SectionTitle"]))
    elements.append(Paragraph(
        "La siguiente tabla presenta los 20 conflictos con el mayor puntaje de riesgo. "
        "Un <b>puntaje de riesgo</b> alto indica que el usuario tiene accesos que le permitirian "
        "ejecutar actividades incompatibles sin supervision, lo cual incrementa el riesgo de "
        "fraude, errores operacionales o incumplimiento normativo.",
        styles["BodyText2"],
    ))

    if critical_conflicts:
        table_header = [
            Paragraph("<b>#</b>", styles["TableHeader"]),
            Paragraph("<b>Usuario SAP</b>", styles["TableHeader"]),
            Paragraph("<b>Nombre</b>", styles["TableHeader"]),
            Paragraph("<b>Regla Vulnerada</b>", styles["TableHeader"]),
            Paragraph("<b>Severidad</b>", styles["TableHeader"]),
            Paragraph("<b>Risk Score</b>", styles["TableHeader"]),
        ]
        table_data = [table_header]

        for i, (conflict, user_id, full_name, rule_name, severity) in enumerate(critical_conflicts, 1):
            sev_val = severity.value if hasattr(severity, 'value') else str(severity)
            table_data.append([
                Paragraph(str(i), styles["TableCellCenter"]),
                Paragraph(str(user_id), styles["TableCell"]),
                Paragraph(str(full_name or "-"), styles["TableCell"]),
                Paragraph(str(rule_name), styles["TableCell"]),
                Paragraph(f"<b>{sev_val}</b>", styles["TableCellCenter"]),
                Paragraph(f"<b>{conflict.risk_score}</b>", styles["TableCellCenter"]),
            ])

        col_widths = [
            content_w * 0.05,  # #
            content_w * 0.14,  # User
            content_w * 0.20,  # Name
            content_w * 0.35,  # Rule
            content_w * 0.12,  # Severity
            content_w * 0.14,  # Score
        ]
        t = Table(table_data, colWidths=col_widths, repeatRows=1)

        # Table styling
        style_cmds = [
            # Header
            ("BACKGROUND", (0, 0), (-1, 0), PRIMARY),
            ("TEXTCOLOR", (0, 0), (-1, 0), WHITE),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, 0), 8.5),
            ("ALIGN", (0, 0), (-1, 0), "CENTER"),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("TOPPADDING", (0, 0), (-1, -1), 5),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
            ("LEFTPADDING", (0, 0), (-1, -1), 4),
            ("RIGHTPADDING", (0, 0), (-1, -1), 4),
            # Grid
            ("GRID", (0, 0), (-1, -1), 0.5, LIGHT_GRAY),
            ("BOX", (0, 0), (-1, -1), 1, PRIMARY),
        ]

        # Alternating row colors
        for i in range(1, len(table_data)):
            if i % 2 == 0:
                style_cmds.append(("BACKGROUND", (0, i), (-1, i), LIGHT_BG))

        # Color-code severity cells
        for i, (conflict, _, _, _, severity) in enumerate(critical_conflicts, 1):
            sev_val = severity.value if hasattr(severity, 'value') else str(severity)
            if sev_val == "HIGH":
                style_cmds.append(("TEXTCOLOR", (4, i), (4, i), RED))
            elif sev_val == "MEDIUM":
                style_cmds.append(("TEXTCOLOR", (4, i), (4, i), ORANGE))
            else:
                style_cmds.append(("TEXTCOLOR", (4, i), (4, i), GREEN))

            # Color-code risk score
            if conflict.risk_score >= 80:
                style_cmds.append(("TEXTCOLOR", (5, i), (5, i), RED))
            elif conflict.risk_score >= 50:
                style_cmds.append(("TEXTCOLOR", (5, i), (5, i), ORANGE))

        t.setStyle(TableStyle(style_cmds))
        elements.append(t)
    else:
        elements.append(Paragraph("<i>No se detectaron conflictos criticos.</i>", styles["BodyText2"]))

    # ── SECTION 5: Top Users with Most Conflicts ──
    section_num += 1
    if top_users:
        elements.append(Spacer(1, 12))
        elements.append(Paragraph(f"{section_num}. USUARIOS CON MAS CONFLICTOS", styles["SectionTitle"]))
        elements.append(Paragraph(
            "Los usuarios listados a continuacion son aquellos que presentan la mayor cantidad "
            "de conflictos de Segregacion de Funciones. Se recomienda revisar de forma prioritaria "
            "los accesos de estos usuarios y evaluar la necesidad de restringir permisos.",
            styles["BodyText2"],
        ))

        user_header = [
            Paragraph("<b>#</b>", styles["TableHeader"]),
            Paragraph("<b>Usuario SAP</b>", styles["TableHeader"]),
            Paragraph("<b>Nombre Completo</b>", styles["TableHeader"]),
            Paragraph("<b>Cantidad de Conflictos</b>", styles["TableHeader"]),
        ]
        user_table_data = [user_header]
        for i, (user_id, full_name, cnt) in enumerate(top_users, 1):
            user_table_data.append([
                Paragraph(str(i), styles["TableCellCenter"]),
                Paragraph(str(user_id), styles["TableCell"]),
                Paragraph(str(full_name or "-"), styles["TableCell"]),
                Paragraph(f"<b>{cnt}</b>", styles["TableCellCenter"]),
            ])

        user_col_widths = [content_w * 0.08, content_w * 0.22, content_w * 0.45, content_w * 0.25]
        ut = Table(user_table_data, colWidths=user_col_widths, repeatRows=1)
        ut.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), PRIMARY),
            ("TEXTCOLOR", (0, 0), (-1, 0), WHITE),
            ("ALIGN", (0, 0), (-1, 0), "CENTER"),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("TOPPADDING", (0, 0), (-1, -1), 5),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
            ("GRID", (0, 0), (-1, -1), 0.5, LIGHT_GRAY),
            ("BOX", (0, 0), (-1, -1), 1, PRIMARY),
        ] + [("BACKGROUND", (0, i), (-1, i), LIGHT_BG) for i in range(2, len(user_table_data), 2)]))
        elements.append(ut)

    # ── SECTION 6: Findings Status ─────────────
    section_num += 1
    if findings:
        elements.append(Spacer(1, 12))
        elements.append(Paragraph(f"{section_num}. ESTADO DE HALLAZGOS", styles["SectionTitle"]))
        elements.append(Paragraph(
            "Los hallazgos son las observaciones formales generadas a partir de los conflictos "
            "detectados. Cada hallazgo tiene un estado que indica su progreso: "
            "<b>ABIERTO</b> (pendiente de revision), <b>EN REVISION</b> (siendo evaluado), "
            "<b>ACEPTADO</b> (reconocido por el responsable), <b>REMEDIADO</b> (corregido) o "
            "<b>CERRADO</b> (verificado y finalizado).",
            styles["BodyText2"],
        ))

        status_labels = {
            "OPEN": "Abierto",
            "IN_REVIEW": "En Revision",
            "ACCEPTED": "Aceptado",
            "REMEDIATED": "Remediado",
            "EXCEPTION_APPROVED": "Excepcion Aprobada",
            "CLOSED": "Cerrado",
        }

        # Donut chart for findings status
        status_data = []
        status_colors_list = []
        status_legend = []
        status_palette = {
            "OPEN": RED, "IN_REVIEW": ORANGE, "ACCEPTED": HexColor("#7B1FA2"),
            "REMEDIATED": ACCENT, "EXCEPTION_APPROVED": HexColor("#FF6F00"), "CLOSED": GREEN,
        }

        for status_key in ["OPEN", "IN_REVIEW", "ACCEPTED", "REMEDIATED", "EXCEPTION_APPROVED", "CLOSED"]:
            cnt = finding_status_counts.get(status_key, 0)
            if cnt > 0:
                status_data.append(cnt)
                status_colors_list.append(status_palette.get(status_key, GRAY))
                status_legend.append((status_palette.get(status_key, GRAY),
                                      f"{status_labels.get(status_key, status_key)}: {cnt}"))

        if status_data:
            finding_drawing = Drawing(content_w, 160)
            fpie = Pie()
            fpie.x = content_w / 2 - 60
            fpie.y = 10
            fpie.width = 120
            fpie.height = 120
            fpie.data = status_data
            fpie.labels = None
            fpie.slices.strokeWidth = 0.5
            fpie.slices.strokeColor = WHITE
            # Donut effect

            for i, color in enumerate(status_colors_list):
                fpie.slices[i].fillColor = color

            finding_drawing.add(fpie)
            finding_drawing.add(Circle(
                fpie.x + (fpie.width / 2),
                fpie.y + (fpie.height / 2),
                (min(fpie.width, fpie.height) / 2) * 0.55,
                fillColor=WHITE,
                strokeColor=WHITE,
            ))

            leg_x = content_w / 2 + 90
            for i, (color, text) in enumerate(status_legend):
                y_pos = 120 - i * 18
                finding_drawing.add(Rect(leg_x, y_pos, 10, 10, fillColor=color, strokeColor=None))
                finding_drawing.add(String(leg_x + 15, y_pos + 1, text,
                                           fontName="Helvetica", fontSize=8.5, fillColor=DARK_TEXT))

            elements.append(finding_drawing)

    # ── SECTION 7: Remediation Plan ────────────
    section_num += 1
    elements.append(Spacer(1, 12))
    elements.append(Paragraph(f"{section_num}. PLAN DE REMEDIACION Y RECOMENDACIONES", styles["SectionTitle"]))
    elements.append(Paragraph(
        "Con base en los hallazgos detectados, se recomienda implementar las siguientes "
        "acciones correctivas para mitigar los riesgos identificados:",
        styles["BodyText2"],
    ))

    recommendations = [
        ["1", "Priorizar conflictos de severidad ALTA",
         "Atender de forma inmediata los conflictos clasificados como de alto riesgo, ya que "
         "representan la mayor exposicion a fraude o errores criticos."],
        ["2", "Revisar accesos de usuarios con mas conflictos",
         "Los usuarios con multiples conflictos deben ser evaluados para determinar si sus "
         "accesos son necesarios o si pueden ser restringidos sin afectar la operacion."],
        ["3", "Implementar controles compensatorios",
         "En casos donde la segregacion total no sea viable, documentar y aprobar controles "
         "compensatorios (revisiones periodicas, aprobaciones duales, monitoreo de logs)."],
        ["4", "Asignar responsables y fechas compromiso",
         "Cada hallazgo debe tener un responsable asignado y una fecha limite de remediacion "
         "para garantizar el seguimiento adecuado."],
        ["5", "Realizar auditoria de seguimiento",
         "Se recomienda programar una auditoria de seguimiento en 30-60 dias para verificar "
         "que las acciones correctivas fueron implementadas correctamente."],
    ]

    rec_header = [
        Paragraph("<b>#</b>", styles["TableHeader"]),
        Paragraph("<b>Recomendacion</b>", styles["TableHeader"]),
        Paragraph("<b>Descripcion</b>", styles["TableHeader"]),
    ]
    rec_data = [rec_header]
    for num, title, desc in recommendations:
        rec_data.append([
            Paragraph(num, styles["TableCellCenter"]),
            Paragraph(f"<b>{title}</b>", styles["TableCell"]),
            Paragraph(desc, styles["TableCell"]),
        ])

    rec_table = Table(rec_data, colWidths=[content_w * 0.06, content_w * 0.30, content_w * 0.64], repeatRows=1)
    rec_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), PRIMARY),
        ("TEXTCOLOR", (0, 0), (-1, 0), WHITE),
        ("ALIGN", (0, 0), (-1, 0), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("TOPPADDING", (0, 0), (-1, -1), 6),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
        ("LEFTPADDING", (0, 0), (-1, -1), 6),
        ("GRID", (0, 0), (-1, -1), 0.5, LIGHT_GRAY),
        ("BOX", (0, 0), (-1, -1), 1, PRIMARY),
    ] + [("BACKGROUND", (0, i), (-1, i), LIGHT_BG) for i in range(2, len(rec_data), 2)]))
    elements.append(rec_table)

    # ── FOOTER NOTE ────────────────────────────
    elements.append(Spacer(1, 20))
    elements.append(HRFlowable(width="100%", thickness=1, color=LIGHT_GRAY, spaceAfter=8))
    elements.append(Paragraph(
        "<b>NOTA:</b> Este reporte es generado automaticamente por el Sistema de Auditoria SoD. "
        "Los datos presentados corresponden al periodo evaluado y reflejan el estado de los accesos "
        "al momento de la importacion de datos SAP. Este documento es de caracter confidencial y "
        "esta dirigido exclusivamente a los responsables de la auditoria.",
        styles["SmallGray"],
    ))

    # ── 4. Build the PDF ───────────────────────
    def add_page_number(canvas_obj, doc):
        canvas_obj.saveState()
        canvas_obj.setFont("Helvetica", 7)
        canvas_obj.setFillColor(GRAY)
        canvas_obj.drawRightString(
            page_w - 40, 25,
            f"Pagina {doc.page} | {audit.company_name} | Reporte Ejecutivo SoD",
        )
        # Header line on every page except first
        if doc.page > 1:
            canvas_obj.setStrokeColor(PRIMARY)
            canvas_obj.setLineWidth(1.5)
            canvas_obj.line(40, page_h - 35, page_w - 40, page_h - 35)
            canvas_obj.setFont("Helvetica-Bold", 8)
            canvas_obj.setFillColor(PRIMARY)
            canvas_obj.drawString(40, page_h - 30, f"AUDITORIA SoD | {audit.name}")
        canvas_obj.restoreState()

    output = BytesIO()
    doc = SimpleDocTemplate(
        output, pagesize=A4,
        leftMargin=50, rightMargin=50, topMargin=45, bottomMargin=45,
        title=f"Reporte Ejecutivo - {audit.name}",
        author="Sistema de Auditoria SoD",
    )
    doc.build(elements, onFirstPage=add_page_number, onLaterPages=add_page_number)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="executive_report_{audit_id}.pdf"'},
    )
